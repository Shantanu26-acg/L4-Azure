from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import tempfile


options = Options()
options.add_argument(f'--user-data-dir={tempfile.mkdtemp()}')
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=options)
#driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(8)

wb = load_workbook("ACG_Common_Workbook.xlsx",data_only=True)

url_login_cred = wb["URL_Login_cred_Tenant"]

url = url_login_cred.cell(row = 2, column = 1).value
driver.get(str(url))

usern = url_login_cred.cell(row = 2, column = 2).value
driver.find_element(By.NAME,'identifier').send_keys(str(usern))

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))

driver.find_element(By.XPATH, "//button[text()='Login']").click()

# Navigate to Serial Number Management -> Download Serial Numbers
try:
    el_sng = EC.element_to_be_clickable((By.XPATH, "//span[text()='Serial Number Management']"))
    WebDriverWait(driver,60).until(el_sng)
    driver.find_element(By.XPATH, "//span[text()='Serial Number Management']").click()
except:
    print("Not found")

#click on sng request history
driver.find_element(By.XPATH, "//li[text()='SNG Request History']").click()

#fetch required sheets
serial_template = wb['SNG_Template']
product = wb["Product"]
start_row=2

for i in range(start_row, serial_template.max_row+1):


    def template_create(packcode):
        el = EC.element_to_be_clickable((By.XPATH,"//span[text()='+ Generate']"))
        WebDriverWait(driver,20).until(el)

        driver.find_element(By.XPATH,"//span[text()='+ Generate']").click()

        try:
            elm2 = EC.element_to_be_clickable((By.XPATH,"//span[text()='Cancel']"))
            WebDriverWait(driver,10).until(elm2)
        except TimeoutException:
            print("SNG download cancel button: Timed out waiting for page to load")
    
        # Product Name - Product Code dropdown      
        driver.find_element(By.XPATH,"//span[text()='Product Name - Product Code']").click()
        el_prod = EC.element_to_be_clickable((By.XPATH,f"//span[text()='{str(prod)}']"))
        WebDriverWait(driver,20).until(el_prod)

        elp = driver.find_element(By.XPATH,f"//span[text()='{str(prod)}']")
        driver.execute_script("arguments[0].scrollIntoView(true)", elp)
        print(prod)
        time.sleep(2)
        driver.find_element(By.XPATH,f"//span[text()='{str(prod)}']").click()
        
        driver.find_element(By.XPATH,"//span[contains(text(),'Primary Pack')]").click()
        driver.find_element(By.XPATH,f"//span[text()='{str(packcode)}']").click()

                   
        driver.find_element(By.XPATH,"//span[text()='Location Name - GLN']").click()
        driver.find_element(By.XPATH,f"//span[contains(text(),'{str(lctn)} -')]").click()     
        
        # Encoding dropdown (URN)
        encoding_type = serial_template.cell(row=i,column=4).value
        driver.find_element(By.XPATH, "//span[contains(text(),'E.g., Digital Link')]").click()
        driver.find_element(By.XPATH, f"//span[contains(text(),'{str(encoding_type)}')]").click()

        qty = serial_template.cell(row=i, column=22).value
        # Serial Number Quantity input
        quantity_input=driver.find_element(By.CSS_SELECTOR, "input[placeholder='Serial Number Quantity']")
        quantity_input.click()
        quantity_input.send_keys(str(qty))

        # File type dropdown (CSV/JSON)
        driver.find_element(By.XPATH, "//span[contains(text(),'CSV/JSON')]").click()
        driver.find_element(By.XPATH, "//span[text()='csv']").click()

        # Submit button
        driver.find_element(By.XPATH, "//button[text()='Submit']").click()

        #to wait until notification is popped up
        el_g = EC.element_to_be_clickable((By.XPATH,"//span[text()='+ Generate']"))
        WebDriverWait(driver,60).until(el_g)

        # Click notification icon (assuming it is a button with role or aria-label, update if needed)
        driver.find_element(By.XPATH,"//span[@class='p-button-icon p-c pi pi-bell']").click()

        try:
            el_number = EC.element_to_be_clickable((By.XPATH,"//div[@class='p-card-content']/a"))
            WebDriverWait(driver,30).until(el_number)
            driver.find_element(By.XPATH, "//div[@class='p-card-content']/a").click()
            print("Serial file got downloaded")
        except:
            print("There is no Serial number request File")
        
        time.sleep(2)

        try:
            #msg = EC.visibility_of_element_located((By.XPATH,"//div[@class='p-card-content']/p"))
            #WebDriverWait(driver,3).until(msg)
            msg1 = driver.find_element(By.XPATH,"//div[@class='p-card-content']/p").get_attribute("innerText")
            print(msg1)
        except:
            print("No messg")

        time.sleep(2)

        try:
            driver.find_element(By.XPATH,"//span[text()='Clear']").click()
        except:
            print("No file found")

        driver.find_element(By.XPATH,"//button[@aria-label='Close']").click()
        
        print("Done for template",str(packcode))
        
    No_of_levels = serial_template.cell(row = i, column = 13).value
    
    actions = ActionChains(driver)

    temp_name = serial_template.cell(row=i, column=1).value
    sng = serial_template.cell(row=i, column=2).value
    prod = str(product[f"B{i}"].value)+"-"+str(product[f"A{i}"].value)
    print(prod) #concating data from product table
    pack_code_1 = str(product[f"Z{i}"].value)+"-"+str(product[f"N{i}"].value)
    lctn = serial_template.cell(row=i, column=6).value

    template_create(pack_code_1)

    if No_of_levels in (2,3,4):
        pack_code_2 = str(product[f"AA{i}"].value)+"-"+str(product[f"Q{i}"].value)
        template_create(pack_code_2)
    
    if No_of_levels in (3,4):
        pack_code_3 = str(product[f"AB{i}"].value)+"-"+str(product[f"T{i}"].value)
        template_create(pack_code_3)
        
    if No_of_levels == 4:
        pack_code_4 = str(product[f"AC{i}"].value)+"-"+str(product[f"W{i}"].value)
        template_create(pack_code_4)

    time.sleep(5) 
    wb.save("ACG_Common_Workbook.xlsx")
