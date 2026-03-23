from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import tempfile

options = Options()
options.add_argument(f'--user-data-dir={tempfile.mkdtemp()}')
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.implicitly_wait(60)

        
wb = load_workbook("ACG_Common_Workbook.xlsx",data_only=True)

url_login_cred = wb["URL_Login_cred_Tenant"]

url = url_login_cred.cell(row = 2, column = 1).value
driver.get(str(url))

#driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[3]/i").click()

usern = url_login_cred.cell(row = 2, column = 2).value
driver.find_element(By.NAME,'identifier').send_keys(str(usern))

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))

driver.find_element(By.XPATH, "//button[text()='Login']").click()

driver.find_element(By.XPATH,"//button[@aria-label='Menu']").click()

#click on serial management
driver.find_element(By.XPATH,"//span[text()='Serial Number Management']").click()
#click on view templates
driver.find_element(By.XPATH,"//li[text()='View Templates']").click()


#fetch template sheet
serial_template = wb['SNG_Template']
product = wb["Product"]
start_row = 2
last_row = (serial_template.max_row)+1

for i in range(start_row, last_row):
    
    def template_create(temp_nam,packcode,colmn):
        timeout=10
        try:
            elm = EC.element_to_be_clickable((By.XPATH,"//span[text()='+ New Template']"))
            WebDriverWait(driver,timeout).until(elm)
        except TimeoutException:
            print("Create template: Timed out waiting for page to load")
        
        time.sleep(5)
        
        driver.find_element(By.XPATH,"//span[text()='+ New Template']").click()
        
        try:
            elm2 = EC.element_to_be_clickable((By.XPATH,"//button[text()='Cancel']"))
            WebDriverWait(driver,timeout).until(elm2)
        except TimeoutException:
            print("SNG cancel button: Timed out waiting for page to load")
        
        #basic details page
        driver.find_element(By.NAME,"templateName").send_keys(temp_nam)
        print("temp_nam:", temp_nam)
        
        driver.find_element(By.XPATH,"//span[text()='Serial Number Generator']").click()
        driver.find_element(By.XPATH,"//span[text()='"+str(sng)+"']").click()
        
        driver.find_element(By.XPATH,"//span[text()='Product Name - Product Code']").click()
        time.sleep(3)
        driver.find_element(By.XPATH,"//span[text()='"+str(prod)+"']").click()
        
        
        driver.find_element(By.XPATH,"//span[text()='Select Packing']").click()
        driver.find_element(By.XPATH,"//span[text()='"+str(packcode)+"']").click()
        print(packcode)

        el = driver.find_element(By.XPATH,"//span[text()='Select BusinessPartner']")
        driver.execute_script("arguments[0].scrollIntoView(true)",el)
                   
        driver.find_element(By.XPATH,"//span[text()='Select BusinessPartner']").click()
        time.sleep(3)
        driver.find_element(By.XPATH,"//span[text()='"+str(bsnptnr)+"']").click()     
                        
        # driver.find_element(By.XPATH,"//*[contains(text(),'"+str(lctn)+"')]").click()

        driver.find_element(
            By.XPATH, f"//label[contains(normalize-space(.), '{lctn.strip()}')]"
        ).click()

        driver.find_element(By.XPATH,"//span[text()='Add']").click()
        
        #click on next page
        driver.find_element(By.XPATH,"//button[text()='Next']").click()
        
        
        #format page
        driver.find_element(By.XPATH,"//div[@id='NumberType']/span[contains(text(),'Select')]").click()
        driver.find_element(By.XPATH,"//span[text()='"+str(num_typ)+"']").click()
        
        if num_typ == 'Numeric':
            driver.find_element(By.XPATH,"//div[@id='generationType']/span[contains(text(),'Select')]").click()
            driver.find_element(By.XPATH,"//span[text()='"+str(gen_typ)+"']").click()
        
        
        driver.find_element(By.ID,"length").send_keys(str(num_len))
        
        #prefix
        #pref = serial_template.cell(row=i, column=10).value
        #driver.find_element(By.XPATH,"//input[@name='prefix']").send_keys(str(pref))
        
        #suffix
        #suff = serial_template.cell(row=i, column=11).value
        #driver.find_element(By.XPATH,"//input[@name='suffix']").send_keys(str(suff))
    
    
        #click on next page
        driver.find_element(By.XPATH,"//button[text()='Next']").click()
        
        
        #pool details
        driver.find_element(By.NAME,"initialPoolSize").send_keys(str(pool_size))
        
        #click on submit
        driver.find_element(By.XPATH,"//button[text()='Submit']").click()
        message = driver.find_element(By.XPATH,"//div[@class='p-toast-detail']").get_attribute("innerText")
        print(message)
        
        time.sleep(3)
        
        if "successfully" in message:
            elem = EC.element_to_be_clickable((By.XPATH,"//span[text()='+ New Template']"))
            WebDriverWait(driver, timeout).until(elem)
        else:
            driver.find_element(By.XPATH,"//button[text()='Cancel']").click()
        
        serial_template.cell(row=i, column=colmn).value=message
        
        print("Done for template",str(packcode))
        
    No_of_levels = serial_template.cell(row = i, column = 13).value

    # for col in ["M", "N", "O", "P", "Q", "Z"]:
    #     print(col, product[f"{col}{i}"].value)


    temp_name = serial_template.cell(row=i, column=1).value
    sng = serial_template.cell(row=i, column=2).value
    prod = str(product[f"B{i}"].value)+"-"+str(product[f"A{i}"].value)
    print(prod) #concating data from product table
    pack_code_1 = str(product[f"Z{i}"].value)+"-"+str(product[f"N{i}"].value)
    print(product[f"N{i}"].value)
    bsnptnr = serial_template.cell(row=i, column=5).value
    lctn = serial_template.cell(row=i, column=6).value
    num_typ = serial_template.cell(row=i, column = 7).value
    gen_typ = serial_template.cell(row=i, column = 8).value
    num_len = serial_template.cell(row=i, column = 9).value
    pool_size = serial_template.cell(row=i, column=12).value

    # template_create((str(temp_name)+"_1"),pack_code_1,18)
    #
    # if No_of_levels in (2,3,4):
    #     pack_code_2 = str(product[f"AA{i}"].value)+"-"+str(product[f"Q{i}"].value)
    #     print("DEBUG: Pack Code 2 =", pack_code_2)
    #     template_create((str(temp_name)+"_2"),pack_code_2,19)
    #
    # if No_of_levels in (3,4):
    #     pack_code_3 = str(product[f"AB{i}"].value)+"-"+str(product[f"T{i}"].value)
    #     print("DEBUG: Pack Code 3 =", pack_code_3)
    #     template_create((str(temp_name)+"_3"),pack_code_3,20)
    #
    # if No_of_levels == 4:
    #     pack_code_4 = str(product[f"AC{i}"].value)+"-"+str(product[f"W{i}"].value)
    #     print("DEBUG: Pack Code 4 =", pack_code_4)
    #     template_create((str(temp_name)+"_4"),pack_code_4,21)

    # now call selenium logic
    print("================================================")
    print(f"Row {i}")
    print("Template Name:", temp_name)
    print("SNG:", sng)
    print("Product:", prod)
    print("Pack Code 1:", pack_code_1)
    print("Business Partner:", bsnptnr)
    print("Location:", lctn)
    print("Number Type:", num_typ)
    print("Generation Type:", gen_typ)
    print("Number Length:", num_len)
    print("Pool Size:", pool_size)
    print("No of Levels:", No_of_levels)
    print("================================================")

    template_create((str(temp_name) + "_1"), pack_code_1, 18)

    if No_of_levels in (2, 3, 4):
        pack_code_2 = str(product[f"AA{i}"].value) + "-" + str(product[f"Q{i}"].value)
        print("Pack Code 2:", pack_code_2)  # outside selenium
        template_create((str(temp_name) + "_2"), pack_code_2, 19)

    if No_of_levels in (3, 4):
        pack_code_3 = str(product[f"AB{i}"].value) + "-" + str(product[f"T{i}"].value)
        print("Pack Code 3:", pack_code_3)
        template_create((str(temp_name) + "_3"), pack_code_3, 20)

    if No_of_levels == 4:
        pack_code_4 = str(product[f"AC{i}"].value) + "-" + str(product[f"W{i}"].value)
        print("Pack Code 4:", pack_code_4)
        template_create((str(temp_name) + "_4"), pack_code_4, 21)

    # wb.save("ACG_Common_Workbook.xlsx")
        
    wb.save("ACG_Common_Workbook.xlsx")