from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time
from datetime import datetime
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
import tempfile
import os

options = Options()
options.add_argument(f'--user-data-dir={tempfile.mkdtemp()}')
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.implicitly_wait(20)

        
wb = load_workbook("ACG_Common_Workbook.xlsx",data_only=True)
typeev = wb["URL_Login_cred_Tenant"]
type_env = typeev.cell(row = 2, column = 4).value


if type_env == "ACG":
    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_ACG"]

if type_env == "Tenant":
    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_Tenant"]

url = url_login_cred.cell(row = 2, column = 1).value
driver.get(str(url))

#driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[3]/i").click()

usern = url_login_cred.cell(row = 2, column = 2).value
driver.find_element(By.NAME,'identifier').send_keys(str(usern))

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))

driver.find_element(By.XPATH, "//button[text()='Login']").click()

time.sleep(5)
user = wb['Users']
start_user = 2
last_user = (user.max_row) +1

driver.find_element(By.XPATH,"//button[@aria-label='Menu']").click()

try:
    UM = EC.element_to_be_clickable((By.XPATH,"//span[text()='User Management']"))
    WebDriverWait(driver,60).until(UM)
    driver.find_element(By.XPATH,"//span[text()='User Management']").click()
except:
    print("Not found")

driver.find_element(By.XPATH,"//li[text()='View Users']").click() 

for i in range(start_user, last_user):

    timeout = 10
    try:
        new_user_button = EC.element_to_be_clickable((By.XPATH,"//button[text()='+ New User']"))
        WebDriverWait(driver,timeout).until(new_user_button)
    except TimeoutException:
        print("Create user page: Timed out waiting for page to load")

    driver.find_element(By.XPATH,"//button[text()='+ New User']").click()
    
    timeout = 10
    try:
        new_user_can = EC.element_to_be_clickable((By.XPATH,"//span[text()='Cancel']"))
        WebDriverWait(driver,timeout).until(new_user_can)
    except TimeoutException:
        print("Create user page_nxt: Timed out waiting for page to load")
    
    
    name = user.cell(row = i, column = 1).value
    val = str(datetime.now()).split(".")
    ext=val[0].replace("-","").replace(":","").replace(" ","")
    driver.find_element(By.NAME,"userName").send_keys(str(name)+ext)
    
    email = user.cell(row = i, column = 2).value
    #driver.find_element(By.NAME,"emailAddress").send_keys(str(name)+ext+"@gmail.com")
    driver.find_element(By.NAME,"emailAddress").send_keys(str(email))
    
    driver.find_element(By.XPATH,"//button[@title='India']").click()

    country_name = user.cell(row=i, column=12).value
    driver.find_element(By.XPATH,f"//span[text()='{country_name}']").click()

    mob = user.cell(row = i, column = 3).value
    driver.find_element(By.ID,"custom-phone-input").send_keys(str(mob))
    
    # select by visible text
    role = user.cell(row = i, column = 4).value
    print(role)
    
    wait = WebDriverWait(driver, timeout=30, poll_frequency=2)
    
    driver.find_element(By.XPATH,"//span[text()='Select role']").click()

    def role_click():
        roleel = driver.find_element(By.XPATH,"//span[text()='"+role+"']")
        driver.execute_script("arguments[0].scrollIntoView(true)",roleel)
        time.sleep(2)
        driver.find_element(By.XPATH,"//span[text()='"+role+"']").click()

    try:
        role_click()
        print("role click executed properly")
    except:
        driver.find_element(By.XPATH,"//span[text()='Select role']").click()
        role_click()

    
    time.sleep(2)
    def location_click():
        loc = driver.find_element(By.XPATH,"//span[text()='"+location+"']")
        driver.execute_script("arguments[0].scrollIntoView(true)",loc)
        time.sleep(2)
        driver.find_element(By.XPATH,"//span[text()='"+location+"']").click()

    if type_env == 'Tenant':
        location = user.cell(row = i, column = 5).value
        print(location)
        driver.find_element(By.XPATH,"//span[text()='Select location']").click()
        try:
            location_click()
            print("location click executed properly")
        except:
            driver.find_element(By.XPATH,"//span[text()='Select location']").click()
            location_click()
        
    #click on next button
    driver.find_element(By.XPATH,"//span[text()='Next']").click()
    print("clicked on next button")

    try:
        #click on submit
        new_user_sub = EC.element_to_be_clickable((By.XPATH, "//span[text()='Submit']"))
        WebDriverWait(driver, timeout).until(new_user_sub)
        driver.find_element(By.XPATH, "//span[text()='Submit']").click()
        print("clicked on submit button")
        #message = driver.find_element(By.XPATH,"//div[@data-status='error' or @data-status='success']").text
        message = driver.find_element(By.XPATH,"//div[@class='p-toast-detail']").get_attribute("innerText")
        print(message)
        if "created" or "successfully" in message:
            new_user = EC.element_to_be_clickable((By.XPATH, "//button[text()='+ New User']"))
            WebDriverWait(driver, timeout).until(new_user) 
        else:
            driver.find_element(By.XPATH,"//span[text()='Cancel']").click()
    except TimeoutException:
        #this loop will execute when error occurs while clicking on next button
        driver.find_element(By.XPATH,"//span[text()='Cancel']").click()

    user.cell(row=i, column=6).value = message
 
    wb.save("ACG_Common_Workbook.xlsx")
    