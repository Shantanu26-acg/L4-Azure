import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import tempfile

options = Options()
options.add_argument(f'--user-data-dir={tempfile.mkdtemp()}')
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.implicitly_wait(5)

        
wb = load_workbook("ACG_Common_Workbook.xlsx",data_only=True)

url_login_cred = wb["URL_Login_cred_Tenant"]

url = url_login_cred.cell(row = 2, column = 1).value
driver.get(str(url))

usern = url_login_cred.cell(row = 2, column = 2).value
driver.find_element(By.NAME,'identifier').send_keys(str(usern))

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))

driver.find_element(By.XPATH, "//button[text()='Login']").click()

try:
    el_sng = EC.element_to_be_clickable((By.XPATH, "//span[text()='Serial Number Management']"))
    WebDriverWait(driver,60).until(el_sng)
    driver.find_element(By.XPATH, "//span[text()='Serial Number Management']").click()
except:
    print("Not found")
#click on view templates
sscc = driver.find_element(By.XPATH,"//li[text()='SSCC Generation Report']")
driver.execute_script("arguments[0].scrollIntoView(true)",sscc) 
driver.find_element(By.XPATH,"//li[text()='SSCC Generation Report']").click()


#fetch template sheet
serial_template = wb['SSCC_template']
start_row = 2
last_row = (serial_template.max_row)+1

for i in range(start_row, last_row):
    
    elm = EC.element_to_be_clickable((By.XPATH,"//span[text()='+ Generate']"))
    WebDriverWait(driver,20).until(elm)
    
    driver.find_element(By.XPATH,"//span[text()='+ Generate']").click()
    
    try:
        elm2 = EC.element_to_be_clickable((By.XPATH,"//span[text()='Cancel']"))
        WebDriverWait(driver,10).until(elm2)
    except TimeoutException:
        print("SSCC download cancel button: Timed out waiting for page to load")
    
    #basic details page
    bsnptnr = serial_template.cell(row=i, column=3).value
    driver.find_element(By.XPATH,"//span[text()='Business Partner']").click()
    el_bs = driver.find_element(By.XPATH,f"//span[contains(text(),'{bsnptnr}-')]")
    driver.execute_script("arguments[0].scrollIntoView(true)", el_bs)
    driver.find_element(By.XPATH,f"//span[contains(text(),'{bsnptnr}-')]").click()    
                    
    lctn = serial_template.cell(row=i, column=7).value
    driver.find_element(By.XPATH,"//span[text()='Physical Location']").click()        
    el_phyl = driver.find_element(By.XPATH,f"//span[text()='{lctn}']")
    driver.execute_script("arguments[0].scrollIntoView(true)",el_phyl)
    driver.find_element(By.XPATH,f"//span[text()='{lctn}']").click()
    
    qty = serial_template.cell(row=i, column=9).value
    driver.find_element(By.ID,"sNoQty").send_keys(str(qty))

    ext = serial_template.cell(row=i, column=5).value
    driver.find_element(By.XPATH,"//input[@id='extensionDigit']").send_keys(str(ext))

    encoding_type = serial_template.cell(row=i,column=10).value
    driver.find_element(By.XPATH, "//span[contains(text(),'Eg.Digital Link')]").click()
    driver.find_element(By.XPATH, f"//span[text()='{str(encoding_type)}']").click()

    # File type dropdown (CSV/JSON)
    driver.find_element(By.XPATH, "//span[contains(text(),'csv/json')]").click()
    driver.find_element(By.XPATH, "//span[text()='csv']").click()
    
    #click on submit
    driver.find_element(By.XPATH,"//span[text()='Submit']").click()
    
    try:
        elm = EC.element_to_be_clickable((By.XPATH,"//span[text()='+ Generate']"))
        WebDriverWait(driver,30).until(elm)
    except:
        driver.find_element(By.XPATH,"//span[text()='Submit']").click()
        messg = driver.find_element(By.XPATH,"//div[@class='p-toast-detail']").get_attribute("innerText")  
        print(messg)
        driver.find_element(By.XPATH,"//span[text()='Cancel']").click()
        serial_template.cell(row=i, column=11).value=messg
    
    wb.save("ACG_Common_Workbook.xlsx")

    time.sleep(3)

