from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import tempfile

options = Options()
options.add_argument(f'--user-data-dir={tempfile.mkdtemp()}')
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.implicitly_wait(60)

        
wb = load_workbook("ACG_Common_Workbook.xlsx",data_only=True)

url_login_cred = wb["URL_Login_cred_Tenant"]

url = url_login_cred.cell(row = 2, column = 1).value
driver.get(str(url))

#driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[3]/i").click()

usern = url_login_cred.cell(row = 2, column = 2).value
driver.find_element(By.NAME,'identifier').send_keys(str(usern))

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))

driver.find_element(By.XPATH, "//button[text()='Login']").click()

driver.find_element(By.XPATH,"//button[@aria-label='Menu']").click()

#click on serial management
driver.find_element(By.XPATH,"//span[text()='Serial Number Management']").click()
#click on view templates
driver.find_element(By.XPATH,"//li[text()='View SSCC Templates']").click()


#fetch template sheet
serial_template = wb['SSCC_template']
start_row = 2
last_row = (serial_template.max_row)+1

for i in range(start_row, last_row):
    
    timeout=10
    try:
        elm = EC.element_to_be_clickable((By.XPATH,"//span[text()='+ New Template']"))
        WebDriverWait(driver,timeout).until(elm)
    except TimeoutException:
        print("Create SSCC template: Timed out waiting for page to load")
    
    time.sleep(3)
    
    driver.find_element(By.XPATH,"//span[text()='+ New Template']").click()
    
    try:
        elm2 = EC.element_to_be_clickable((By.XPATH,"//button[text()='Cancel']"))
        WebDriverWait(driver,timeout).until(elm2)
    except TimeoutException:
        print("SNG cancel button: Timed out waiting for page to load")
    
    #basic details page
    temp_name = serial_template.cell(row=i, column=1).value
    driver.find_element(By.NAME,"templateName").send_keys(str(temp_name))
    
    
    #start_rng = serial_template.cell(row=i, column=2).value
    #driver.find_element(By.ID,"startRange").send_keys(str(start_rng))
    
    bsnptnr = serial_template.cell(row=i, column=3).value
    driver.find_element(By.XPATH,"//span[text()='Select Business Partner (GCP)']").click()
    elem = driver.find_element(By.XPATH,"//span[text()='"+str(bsnptnr)+"']")
    driver.execute_script("arguments[0].scrollIntoView(true)",elem)
    driver.find_element(By.XPATH,"//span[text()='"+str(bsnptnr)+"']").click()
    
    sng = serial_template.cell(row=i, column=4).value
    driver.find_element(By.XPATH,"//span[text()='Select Serial Number Generator']").click()
    driver.find_element(By.XPATH,"//span[text()='"+str(sng)+"']").click()
    
    ext = serial_template.cell(row=i, column=5).value
    driver.find_element(By.XPATH,"//input[@id='"+str(ext)+"']").click()
    
    el = driver.find_element(By.XPATH,"//span[text()='Add']")
    driver.execute_script("arguments[0].scrollIntoView(true)",el)
               
    bsnptnr_rec = serial_template.cell(row=i, column=6).value
    driver.find_element(By.XPATH,"//div[@id='businessPartnerName']/span[text()='Select Business']").click()
    time.sleep(3)
    driver.find_element(By.XPATH,"//li/span[text()='"+str(bsnptnr_rec)+"']").click()     
                    
                    
    lctn = serial_template.cell(row=i, column=7).value
    # driver.find_element(By.XPATH,"//*[contains(text(),'"+str(lctn)+"')]").click()

    driver.find_element(
        By.XPATH, f"//label[contains(normalize-space(.), '{lctn.strip()}')]"
    ).click()

    driver.find_element(By.XPATH,"//span[text()='Add']").click()
    
    
    #click on submit
    driver.find_element(By.XPATH,"//button[text()='Submit']").click()
    message = driver.find_element(By.XPATH,"//div[@class='p-toast-detail']").get_attribute("innerText")
    print(message)
    
    time.sleep(3)
    
    if "successfully" in message:
        elem = EC.element_to_be_clickable((By.XPATH,"//span[text()='+ New Template']"))
        WebDriverWait(driver, timeout).until(elem)
    else:
        driver.find_element(By.XPATH,"//button[@type='button']").click()
        driver.find_element(By.XPATH,"//button[text()='Cancel']").click()
    
    '''try:
        elem = EC.element_to_be_clickable((By.XPATH,"//b[text()='+ New Template']"))
        WebDriverWait(driver, timeout).until(elem)
    except TimeoutException:
        driver.find_element(By.XPATH,"//button[text()='Cancel']").click()'''
     
    
    serial_template.cell(row=i, column=8).value=message
    wb.save("ACG_Common_Workbook.xlsx")