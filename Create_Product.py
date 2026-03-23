import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import tempfile
import os

options = Options()
options.add_argument(f'--user-data-dir={tempfile.mkdtemp()}')
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=options)

driver.maximize_window()
driver.implicitly_wait(60)

workbook = load_workbook("ACG_Common_Workbook.xlsx",data_only=True)
sheet = workbook["Product"]
url_login_cred = workbook["URL_Login_cred_Tenant"]

start_row = 2
#time.sleep(3)

# url = url_login_cred.cell(row = 2, column = 1).value
# driver.get(str(url))

# usern = url_login_cred.cell(row = 2, column = 2).value
# driver.find_element(By.NAME,'identifier').send_keys(str(usern))

# passw = url_login_cred.cell(row = 2, column = 3).value
# driver.find_element(By.NAME,'password').send_keys(str(passw))

# driver.find_element(By.XPATH, "//button[text()='Login']").click()

# driver.find_element(By.XPATH,"//button[@aria-label='Menu']").click()

# # Navigate to the Add Product page
# EC.visibility_of_element_located((By.XPATH, "//span[text()= 'Master Data']"))

# driver.find_element(By.XPATH, "//span[text()= 'Master Data']").click()
# driver.find_element(By.XPATH,"//li[text()= 'Product master data']").click()

# # Loop through all the rows in the Excel sheet
# for i in range(start_row, sheet.max_row + 1):

#     # Click 'Add Product' button at the beginning of each loop
#     timeout = 10
#     try:
#         new_prd = EC.element_to_be_clickable((By.XPATH,"//span[text()= '+ New Product']"))
#         WebDriverWait(driver,timeout).until(new_prd)
#     except TimeoutException:
#         print("Create Product page: Timed out waiting for page to load")

#     driver.find_element(By.XPATH,"//span[text()= '+ New Product']").click()


#     ProductIdentifier = sheet.cell(row = i, column = 1).value
#     productName = sheet.cell(row = i, column = 2).value
#     productDescription = sheet.cell(row = i, column = 3).value
#     manufacturer = sheet.cell(row = i, column = 4).value
#     manufacturername = sheet.cell(row = i, column = 5).value
#     GCP = sheet.cell(row = i, column = 6).value
#     ProductIdentifier2 = sheet.cell(row = i, column = 7).value
#     No_of_levels = sheet.cell(row = i, column = 8).value
#     genericName = sheet.cell(row = i, column = 9).value
#     MinTemp = sheet.cell(row = i, column = 10).value
#     weight = sheet.cell(row = i, column = 11).value
#     strength = sheet.cell(row = i, column = 12).value

#     #productIdentifierType = driver.find_element(By.XPATH, "//select[@name= 'productIdentifierType']")
#     #Select(productIdentifierType).select_by_index(2)

#     driver.find_element(By.XPATH,"//span[text()='Select product identifier type']").click()
#     checkbox2 = driver.find_element(By.XPATH,"//span[text()='Product Code']")
#     driver.execute_script("arguments[0].scrollIntoView(true)", checkbox2)
#     driver.find_element(By.XPATH,"//span[text()='Product Code']").click()

#     checkbox = driver.find_element(By.XPATH, "//input[@name='productIdentifier']")
#     driver.execute_script("arguments[0].scrollIntoView(true)", checkbox)
#     # Fill in the Product details
#     driver.find_element(By.XPATH, "//input[@name='productIdentifier']").send_keys(ProductIdentifier)

#     driver.find_element(By.XPATH, "//input[@name='productName']").send_keys(productName)

#     driver.find_element(By.XPATH, "//input[@name='productDescription']").send_keys(productDescription)

#     checkb = driver.find_element(By.XPATH, "//span[text()='Select manufacturer']")
#     driver.execute_script("arguments[0].scrollIntoView(true)", checkb)
#     driver.find_element(By.XPATH, "//span[text()='Select manufacturer']").click()

#     checkbox1 = driver.find_element(By.XPATH, f"//span[text()='{str(manufacturer)}']")
#     driver.execute_script("arguments[0].scrollIntoView(true)", checkbox1)
#     time.sleep(2)
#     driver.find_element(By.XPATH,f"//span[text()='{str(manufacturer)}']").click()

#     if manufacturer == 'Others':
#         driver.find_element(By.XPATH, "//input[@name='manufacturerName']").send_keys(manufacturername)
#         driver.find_element(By.XPATH, "//input[@name='manufacturerOtherGCP']").send_keys(GCP)

#     time.sleep(3)

#     # Next to Packaging Details
#     nextq = driver.find_element(By.XPATH, "//span[text()= 'Next']")
#     driver.execute_script("arguments[0].scrollIntoView(true)", nextq)
#     driver.find_element(By.XPATH, "//span[text()= 'Next']").click()

#     timeout = 10
#     try:
#         new_prod_load = EC.element_to_be_clickable((By.XPATH,"//span[text()= 'Cancel']"))
#         WebDriverWait(driver,timeout).until(new_prod_load)
#     except TimeoutException:
#         print("Create Product page: Timed out waiting for page to load")

#     time.sleep(3)
    
#     def packaging(pkl,indicator,item_ref,colnno,higher,lvl,qty):

#         print(pkl,indicator,item_ref)

#         time.sleep(1)

#         driver.find_element(By.XPATH, "//input[@name='packagingProductIdentifier']").send_keys(item_ref)

#         time.sleep(1)

#         driver.find_element(By.XPATH, "//span[text()='Select Packaging Level']").click()
#         elem = driver.find_element(By.XPATH,"//span[text()='"+str(pkl)+"']")
#         driver.execute_script("arguments[0].scrollIntoView(true)",elem)
#         driver.find_element(By.XPATH,"//span[text()='"+str(pkl)+"']").click()

#         driver.find_element(By.XPATH, "//input[@name='packagingLevelIndicator']").send_keys(indicator)


#         driver.find_element(By.XPATH, "//span[text()='Select Packaging Code Type']").click()
#         time.sleep(1)
#         elem = driver.find_element(By.XPATH,"//*[@id='dropdownItem_0']/span")
#         driver.execute_script("arguments[0].scrollIntoView(true)",elem)
#         #driver.find_element(By.XPATH,"//span[text()='GTIN-14']").click()
#         driver.find_element(By.XPATH,"//*[@id='dropdownItem_0']/span").click()

#         if higher == 'yes':
#             driver.find_element(By.XPATH,"//span[text()='Yes']").click()
#             driver.find_element(By.XPATH,"//span[text()='Select Child Packaging Level']").click()

#             if lvl == 'yes_2':
#                 driver.execute_script("arguments[0].scrollIntoView(true)",driver.find_element(By.XPATH,"//li[@id='dropdownItem_0']"))
#                 driver.find_element(By.XPATH,"//li[@id='dropdownItem_0']").click()
#             if lvl == 'yes_3':
#                 driver.execute_script("arguments[0].scrollIntoView(true)",driver.find_element(By.XPATH,"//li[@id='dropdownItem_1']"))
#                 driver.find_element(By.XPATH,"//li[@id='dropdownItem_1']").click()
#             if lvl == 'yes_4':
#                 driver.execute_script("arguments[0].scrollIntoView(true)",driver.find_element(By.XPATH,"//li[@id='dropdownItem_2']"))
#                 driver.find_element(By.XPATH,"//li[@id='dropdownItem_2']").click()

#             driver.find_element(By.ID,"maxQuantity_0").send_keys(str(qty))

#         gtin = driver.find_element(By.XPATH,"//input[@placeholder='LevelIndicator + GCP + Item Reference + Check Digit']").get_attribute("value")
#         print(gtin)

#         driver.find_element(By.XPATH, "//span[text()= 'Add']").click()

#         sheet.cell(row=i, column=colnno).value = gtin


#     #packaging details
#     pk_level = sheet.cell(row=i,column = 14).value
#     indicator = sheet.cell(row=i,column = 15).value
#     item_ref = sheet.cell(row=i,column = 16).value
#     packaging(pk_level,indicator,item_ref,26,"no","NA",0)


#     if No_of_levels in (2,3,4):
#         pk_level = sheet.cell(row=i,column = 17).value
#         indicator = sheet.cell(row=i,column = 18).value
#         item_ref = sheet.cell(row=i,column = 19).value
#         maxqty = sheet.cell(row=i,column = 10).value
#         packaging(pk_level,indicator,item_ref,27,"yes","yes_2",maxqty)
#         print("second level executed")


#     if No_of_levels in (3,4):
#         pk_level = sheet.cell(row=i,column = 20).value
#         indicator = sheet.cell(row=i,column = 21).value
#         item_ref = sheet.cell(row=i,column = 22).value
#         maxqty = sheet.cell(row=i,column = 11).value
#         packaging(pk_level,indicator,item_ref,28,"yes","yes_3",maxqty)
#         print("third level executed")

#     if No_of_levels == 4:
#         pk_level = sheet.cell(row=i,column = 23).value
#         indicator = sheet.cell(row=i,column = 24).value
#         item_ref = sheet.cell(row=i,column = 25).value
#         maxqty = sheet.cell(row=i,column = 12).value
#         packaging(pk_level,indicator,item_ref,29,"yes","yes_4",maxqty)
#         print("fourth level executed")

#     '''
#     # Next to Other Details
#     driver.find_element(By.XPATH, "//button[text()= 'Next']").click()


#     driver.find_element(By.XPATH, "//input[@placeholder='Enter generic name']").send_keys(genericName)

#     driver.find_element(By.XPATH, "//input[@placeholder='Enter min temperature']").send_keys(MinTemp)


#     # Next to Regulatory Details
#     driver.find_element(By.XPATH, "//button[text()= 'Next']").click()


#     driver.find_element(By.XPATH, "//b[text()= '+ New Regulation']").click()


#     countryselect = driver.find_element(By.XPATH, "//select[@name= 'country']")
#     Select(countryselect).select_by_index(1)


#     selectRegulation = driver.find_element(By.XPATH, "//select[@name= 'regulation']")
#     Select(selectRegulation).select_by_index(1)


#     driver.find_element(By.XPATH, "//input[@placeholder= 'Enter weight (gm)']").send_keys(weight)

#     driver.find_element(By.XPATH, "//input[@name='strength (mg)']").send_keys(strength)


#     driver.find_element(By.XPATH, "//button[text()= 'Accept']").click()

#     # Next to Custom Details
#     driver.find_element(By.XPATH, "//button[text()= 'Next']").click()
#     '''

#     # Submit Details
#     driver.find_element(By.XPATH, "//button[text()= 'Submit']").click()
#     print("clicked on submit")
#     message = driver.find_element(By.XPATH,"//div[@class='p-toast-detail']").get_attribute("innerText")#text
#     print(message)

#     if "successfully" in message:
#         elm = EC.element_to_be_clickable((By.XPATH,"//span[text()= '+ New Product']"))
#         WebDriverWait(driver,3).until(elm)
#     else:
#         driver.find_element(By.XPATH,"//span[text()='Cancel']").click()

#     '''
#     #timeout1=3
#     try:
#     elm = EC.element_to_be_clickable((By.XPATH,"//span[text()= '+ New Product']"))
#     WebDriverWait(driver,3).until(elm)
#     except TimeoutException:
#     driver.find_element(By.XPATH,"//button[text()='Cancel']").click()'''

#     sheet.cell(row=i, column=35).value=message
#     workbook.save("ACG_Common_Workbook.xlsx")

#     time.sleep(2)




url = url_login_cred.cell(row = 2, column = 1).value
driver.get(str(url))
time.sleep(3)

usern = url_login_cred.cell(row = 2, column = 2).value
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'identifier')))
driver.find_element(By.NAME,'identifier').send_keys(str(usern))
time.sleep(1)

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))
time.sleep(1)

driver.find_element(By.XPATH, "//button[text()='Login']").click()
WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//button[@aria-label='Menu']")))
time.sleep(2)

driver.find_element(By.XPATH,"//button[@aria-label='Menu']").click()
time.sleep(2)

# Navigate to the Add Product page
WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//span[text()= 'Master Data']"))).click()
time.sleep(2)
WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,"//li[text()= 'Product master data']"))).click()
time.sleep(3)


# Loop through all the rows in the Excel sheet
for i in range(start_row, sheet.max_row + 1):

    try:
        new_prd = EC.element_to_be_clickable((By.XPATH,"//span[text()= '+ New Product']"))
        WebDriverWait(driver,10).until(new_prd)
    except TimeoutException:
        print("Create Product page: Timed out waiting for page to load")

    driver.find_element(By.XPATH,"//span[text()= '+ New Product']").click()
    time.sleep(3)

    ProductIdentifier = sheet.cell(row = i, column = 1).value
    productName = sheet.cell(row = i, column = 2).value
    productDescription = sheet.cell(row = i, column = 3).value
    manufacturer = sheet.cell(row = i, column = 4).value
    manufacturername = sheet.cell(row = i, column = 5).value
    GCP = sheet.cell(row = i, column = 6).value
    ProductIdentifier2 = sheet.cell(row = i, column = 7).value
    No_of_levels = sheet.cell(row = i, column = 8).value
    genericName = sheet.cell(row = i, column = 9).value
    MinTemp = sheet.cell(row = i, column = 10).value
    weight = sheet.cell(row = i, column = 11).value
    strength = sheet.cell(row = i, column = 12).value

    driver.find_element(By.XPATH,"//span[text()='Select product identifier type']").click()
    time.sleep(1)
    checkbox2 = driver.find_element(By.XPATH,"//span[text()='Product Code']")
    driver.execute_script("arguments[0].scrollIntoView(true)", checkbox2)
    checkbox2.click()
    time.sleep(1)

    checkbox = driver.find_element(By.XPATH, "//input[@name='productIdentifier']")
    driver.execute_script("arguments[0].scrollIntoView(true)", checkbox)
    checkbox.send_keys(ProductIdentifier)
    time.sleep(1)

    driver.find_element(By.XPATH, "//input[@name='productName']").send_keys(productName)
    driver.find_element(By.XPATH, "//input[@name='productDescription']").send_keys(productDescription)
    time.sleep(1)

    checkb = driver.find_element(By.XPATH, "//span[text()='Select manufacturer']")
    driver.execute_script("arguments[0].scrollIntoView(true)", checkb)
    checkb.click()
    time.sleep(2)

    checkbox1 = driver.find_element(By.XPATH, f"//span[text()='{str(manufacturer)}']")
    driver.execute_script("arguments[0].scrollIntoView(true);", checkbox1)
    checkbox1.click()
    time.sleep(1)

    if manufacturer == 'Others':
        driver.find_element(By.XPATH, "//input[@name='manufacturerName']").send_keys(manufacturername)
        driver.find_element(By.XPATH, "//input[@name='manufacturerOtherGCP']").send_keys(GCP)
        time.sleep(1)

    # Next to Packaging Details
    nextq = driver.find_element(By.XPATH, "//span[text()= 'Next']")
    driver.execute_script("arguments[0].scrollIntoView(true)", nextq)
    nextq.click()
    time.sleep(3)

    try:
        new_prod_load = EC.element_to_be_clickable((By.XPATH,"//span[text()= 'Cancel']"))
        WebDriverWait(driver,10).until(new_prod_load)
    except TimeoutException:
        print("Create Product page: Timed out waiting for page to load")

    time.sleep(2)

    def packaging(pkl,indicator,item_ref,colnno,higher,lvl,qty):

        print(pkl,indicator,item_ref)
        time.sleep(1)

        driver.find_element(By.XPATH, "//input[@name='packagingProductIdentifier']").send_keys(item_ref)
        time.sleep(1)

        driver.find_element(By.XPATH, "//span[text()='Select Packaging Level']").click()
        elem = driver.find_element(By.XPATH,"//span[text()='"+str(pkl)+"']")
        driver.execute_script("arguments[0].scrollIntoView(true)",elem)
        elem.click()
        time.sleep(1)

        driver.find_element(By.XPATH, "//input[@name='packagingLevelIndicator']").send_keys(indicator)
        time.sleep(1)

        driver.find_element(By.XPATH, "//span[text()='Select Packaging Code Type']").click()
        time.sleep(1)
        elem = driver.find_element(By.XPATH,"//*[@id='dropdownItem_0']/span")
        driver.execute_script("arguments[0].scrollIntoView(true)",elem)
        elem.click()
        time.sleep(1)

        if higher == 'yes':
            driver.find_element(By.XPATH,"//span[text()='Yes']").click()
            driver.find_element(By.XPATH,"//span[text()='Select Child Packaging Level']").click()
            time.sleep(1)

            if lvl == 'yes_2':
                driver.execute_script("arguments[0].scrollIntoView(true)",driver.find_element(By.XPATH,"//li[@id='dropdownItem_0']"))
                driver.find_element(By.XPATH,"//li[@id='dropdownItem_0']").click()
            if lvl == 'yes_3':
                driver.execute_script("arguments[0].scrollIntoView(true)",driver.find_element(By.XPATH,"//li[@id='dropdownItem_1']"))
                driver.find_element(By.XPATH,"//li[@id='dropdownItem_1']").click()
            if lvl == 'yes_4':
                driver.execute_script("arguments[0].scrollIntoView(true)",driver.find_element(By.XPATH,"//li[@id='dropdownItem_2']"))
                driver.find_element(By.XPATH,"//li[@id='dropdownItem_2']").click()

            time.sleep(1)
            driver.find_element(By.ID,"maxQuantity_0").send_keys(str(qty))
            time.sleep(1)

        gtin = driver.find_element(By.XPATH,"//input[@placeholder='LevelIndicator + GCP + Item Reference + Check Digit']").get_attribute("value")
        print(gtin)

        driver.find_element(By.XPATH, "//span[text()= 'Add']").click()
        time.sleep(2)

        sheet.cell(row=i, column=colnno).value = gtin

    # Level 1
    pk_level = sheet.cell(row=i,column = 14).value
    indicator = sheet.cell(row=i,column = 15).value
    item_ref = sheet.cell(row=i,column = 16).value
    packaging(pk_level,indicator,item_ref,26,"no","NA",0)

    if No_of_levels in (2,3,4):
        pk_level = sheet.cell(row=i,column = 17).value
        indicator = sheet.cell(row=i,column = 18).value
        item_ref = sheet.cell(row=i,column = 19).value
        maxqty = sheet.cell(row=i,column = 10).value
        packaging(pk_level,indicator,item_ref,27,"yes","yes_2",maxqty)
        print("second level executed")

    if No_of_levels in (3,4):
        pk_level = sheet.cell(row=i,column = 20).value
        indicator = sheet.cell(row=i,column = 21).value
        item_ref = sheet.cell(row=i,column = 22).value
        maxqty = sheet.cell(row=i,column = 11).value
        packaging(pk_level,indicator,item_ref,28,"yes","yes_3",maxqty)
        print("third level executed")

    if No_of_levels == 4:
        pk_level = sheet.cell(row=i,column = 23).value
        indicator = sheet.cell(row=i,column = 24).value
        item_ref = sheet.cell(row=i,column = 25).value
        maxqty = sheet.cell(row=i,column = 12).value
        packaging(pk_level,indicator,item_ref,29,"yes","yes_4",maxqty)
        print("fourth level executed")

    # Submit
    driver.find_element(By.XPATH, "//button[text()= 'Submit']").click()
    time.sleep(3)
    print("clicked on submit")

    message = driver.find_element(By.XPATH,"//div[@class='p-toast-detail']").get_attribute("innerText")
    print(message)

    if "successfully" in message:
        elm = EC.element_to_be_clickable((By.XPATH,"//span[text()= '+ New Product']"))
        WebDriverWait(driver,5).until(elm)
    else:
        driver.find_element(By.XPATH,"//span[text()='Cancel']").click()

    sheet.cell(row=i, column=35).value=message
    workbook.save("ACG_Common_Workbook.xlsx")
    time.sleep(2)
