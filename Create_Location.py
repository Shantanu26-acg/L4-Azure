import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import tempfile


options = Options()
options.add_argument(f'--user-data-dir={tempfile.mkdtemp()}')
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.implicitly_wait(60)

workbook = load_workbook("ACG_Common_Workbook.xlsx",data_only=True)
sheet = workbook["partner"]
url_login_cred = workbook["URL_Login_cred_Tenant"]

start_row = 2

time.sleep(3)

# url = url_login_cred.cell(row = 2, column = 1).value
# driver.get(str(url))

# #driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[3]/i").click()

# usern = url_login_cred.cell(row = 2, column = 2).value
# driver.find_element(By.NAME,'identifier').send_keys(str(usern))

# passw = url_login_cred.cell(row = 2, column = 3).value
# driver.find_element(By.NAME,'password').send_keys(str(passw))

# driver.find_element(By.XPATH, "//button[text()='Login']").click()

# driver.find_element(By.XPATH,"//button[@aria-label='Menu']").click()

# #click on master data and loctaion master data
# driver.find_element(By.XPATH,"//span[text()= 'Master Data']").click()
# driver.find_element(By.XPATH,"//li[text()= 'Location master data']").click()


# # Loop through all the rows in the Excel sheet
# for i in range(start_row, sheet.max_row + 1):  # Adjusting for 0-indexing
    
#     # Click 'New Location' button at the beginning of each loop
#     timeout = 10
#     try:
#         new_role = EC.element_to_be_clickable((By.XPATH,"//span[text()= '+ New Location']"))
#         WebDriverWait(driver,timeout).until(new_role)
#     except TimeoutException:
#         print("Create location page: Timed out waiting for page to load")
        
#     time.sleep(5)
    
#     driver.find_element(By.XPATH,"//span[text()= '+ New Location']").click()


#     location_name = sheet.cell(row = i, column = 1).value
#     location_id = sheet.cell(row = i, column = 2).value
#     state = sheet.cell(row = i, column = 3).value
#     city = sheet.cell(row = i, column = 4).value
#     address = sheet.cell(row = i, column = 5).value
#     postal_code = sheet.cell(row = i, column = 6).value
#     contact_person = sheet.cell(row = i, column = 7).value
#     email_id = sheet.cell(row = i, column = 8).value
#     phone_number = sheet.cell(row = i, column = 9).value
#     website = sheet.cell(row = i, column = 10).value
#     entity = sheet.cell(row = i, column = 11).value
#     bus_entity = sheet.cell(row = i, column = 12).value
#     loc_identifier_ty = sheet.cell(row = i, column = 13).value
#     loc_number = sheet.cell(row=i, column=14).value
#     countryyy = sheet.cell(row=i, column=15).value

#     # Fill in the location details
#     locname = driver.find_element(By.XPATH, "//input[@placeholder='Enter location name']")
#     locname.send_keys(location_name)

#     driver.find_element(By.XPATH, "//span[text()='Select location type']").click()
#     driver.find_element(By.XPATH,"//span[text()='Physical Site']").click()
    
#     el = driver.find_element(By.XPATH, "//span[text()='Select entity type']")
#     driver.execute_script("arguments[0].scrollIntoView(true);", el)
    
#     driver.find_element(By.XPATH, "//span[text()='Select entity type']").click()
#     el_e = driver.find_element(By.XPATH,f"//span[text()='{entity}']")
#     driver.execute_script("arguments[0].scrollIntoView(true);", el_e)
#     driver.find_element(By.XPATH,f"//span[text()='{entity}']").click()

#     el = driver.find_element(By.XPATH, "//span[text()='Select business entity']")
#     driver.execute_script("arguments[0].scrollIntoView(true);", el)

#     driver.find_element(By.XPATH, "//span[text()='Select business entity']").click()
#     time.sleep(2)
#     driver.find_element(By.XPATH,f"//span[text()='{bus_entity}']").click()
    
#     element = driver.find_element(By.XPATH, "//span[text()='Identifier type']")
#     driver.execute_script("arguments[0].scrollIntoView(true);", element)
    
#     time.sleep(2)
    
#     driver.find_element(By.XPATH, "//span[text()='Identifier type']").click()
#     driver.find_element(By.XPATH, f"//span[text()='{loc_identifier_ty}']").click()

#     identifier = driver.find_element(By.NAME,"locationNumber")
#     identifier.clear()
#     identifier.send_keys(loc_number)
    
#     driver.find_element(By.XPATH,"//input[@id='extensionDigit']").click()

#     driver.find_element(By.XPATH, "//button[text()='Next']").click()
#     print("clicked on next button")

    
#     timeout = 10
#     try:
#         next_page = EC.element_to_be_clickable((By.XPATH,"//span[text()='Select country']"))
#         WebDriverWait(driver,timeout).until(next_page)
#     except TimeoutException:
#         print("Create location page: Timed out waiting for page to load")
    
#     # Fill in the address details
#     driver.find_element(By.XPATH, "//span[text()='Select country']").click()
#     #element = driver.find_element(By.XPATH, f"//span[text()='{str(countryyy)}']")
#     #driver.execute_script("arguments[0].scrollIntoView(true);", element)
#     time.sleep(2)
#     #actions = ActionChains(driver)
#     driver.find_element(By.XPATH,f"(//span[text()='{str(countryyy)}'])[2]").click()


#     state_element = driver.find_element(By.XPATH, "//input[@name='state']")
#     state_element.send_keys(state)


#     city_element = driver.find_element(By.XPATH, "//input[@name='city']")
#     city_element.send_keys(city)


#     address_element = driver.find_element(By.XPATH, "//input[@name='address']")
#     address_element.send_keys(address)


#     pcode = driver.find_element(By.XPATH, "//input[@name='postalCode']")
#     pcode.send_keys(postal_code)


#     # Scroll into view and fill in contact details
#     element = driver.find_element(By.XPATH, "//input[@name='contactPersonName']")
#     driver.execute_script("arguments[0].scrollIntoView(true);", element)


#     conper = driver.find_element(By.XPATH, "//input[@placeholder='Enter name']")
#     conper.send_keys(contact_person)


#     email_element = driver.find_element(By.XPATH, "//input[@placeholder='Enter email']")
#     email_element.send_keys(email_id)


#     driver.find_element(By.XPATH,"//button[@title='India']").click()

#     element = driver.find_element(By.XPATH, "//span[text()='"+str(countryyy)+"']")
#     driver.execute_script("arguments[0].scrollIntoView(true);", element)
#     driver.find_element(By.XPATH,f"//span[text()='{countryyy}']").click()

#     driver.find_element(By.ID,"custom-phone-input").send_keys(str(phone_number))


#     website_element = driver.find_element(By.XPATH, "//input[@placeholder='Enter website']")
#     website_element.send_keys(website)


#     # Submit the form
#     submit = driver.find_element(By.XPATH, "//button[text()='Submit']")
#     submit.click()
#     message = driver.find_element(By.XPATH,"//div[@class='p-toast-detail']").get_attribute("innerText")
#     print(message)
    
    
#     if "Successfully" in message:
#         elm = EC.element_to_be_clickable((By.XPATH,"//span[text()= '+ New Location']"))
#         WebDriverWait(driver,3).until(elm)
#     #else:
#     #    driver.find_element(By.XPATH,"//button[text()='Cancel']").click()
    
#     '''
#     timeout1=3
#     try:
#         elm = EC.element_to_be_clickable((By.XPATH,"//b[text()= '+ New Location']"))
#         WebDriverWait(driver,timeout1).until(elm)
#     except TimeoutException:
#         driver.find_element(By.XPATH,"//button[text()='Cancel']").click()'''
    
#     sheet.cell(row=i, column=21).value=message
#     workbook.save("ACG_Common_Workbook.xlsx")


url = url_login_cred.cell(row = 2, column = 1).value
driver.get(str(url))
time.sleep(3)

usern = url_login_cred.cell(row = 2, column = 2).value
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'identifier')))
driver.find_element(By.NAME,'identifier').send_keys(str(usern))
time.sleep(1)

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))
time.sleep(1)

driver.find_element(By.XPATH, "//button[text()='Login']").click()
WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"//button[@aria-label='Menu']")))
time.sleep(2)

driver.find_element(By.XPATH,"//button[@aria-label='Menu']").click()
time.sleep(2)

# click on master data and location master data
WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,"//span[text()='Master Data']"))).click()
time.sleep(2)
WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,"//li[text()='Location master data']"))).click()
time.sleep(5)


# Loop through all the rows in the Excel sheet
for i in range(start_row, sheet.max_row + 1):  
    
    # Wait for 'New Location' button
    try:
        new_role = EC.element_to_be_clickable((By.XPATH,"//span[text()= '+ New Location']"))
        WebDriverWait(driver,10).until(new_role)
    except TimeoutException:
        print("Create location page: Timed out waiting for page to load")
        
    time.sleep(3)
    
    driver.find_element(By.XPATH,"//span[text()= '+ New Location']").click()
    time.sleep(3)

    # Read Excel values
    location_name = sheet.cell(row = i, column = 1).value
    location_id = sheet.cell(row = i, column = 2).value
    state = sheet.cell(row = i, column = 3).value
    city = sheet.cell(row = i, column = 4).value
    address = sheet.cell(row = i, column = 5).value
    postal_code = sheet.cell(row = i, column = 6).value
    contact_person = sheet.cell(row = i, column = 7).value
    email_id = sheet.cell(row = i, column = 8).value
    phone_number = sheet.cell(row = i, column = 9).value
    website = sheet.cell(row = i, column = 10).value
    entity = sheet.cell(row = i, column = 11).value
    bus_entity = sheet.cell(row = i, column = 12).value
    loc_identifier_ty = sheet.cell(row = i, column = 13).value
    loc_number = sheet.cell(row=i, column=14).value
    countryyy = sheet.cell(row=i, column=15).value

    # Fill in the location details
    locname = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Enter location name']"))
    )
    locname.send_keys(location_name)
    time.sleep(1)

    driver.find_element(By.XPATH, "//span[text()='Select location type']").click()
    time.sleep(1)
    driver.find_element(By.XPATH,"//span[text()='Physical Site']").click()
    time.sleep(1)
    
    el = driver.find_element(By.XPATH, "//span[text()='Select entity type']")
    driver.execute_script("arguments[0].scrollIntoView(true);", el)
    time.sleep(1)
    
    driver.find_element(By.XPATH, "//span[text()='Select entity type']").click()
    time.sleep(2)
    el_e = driver.find_element(By.XPATH,f"//span[text()='{entity}']")
    driver.execute_script("arguments[0].scrollIntoView(true);", el_e)
    el_e.click()
    time.sleep(1)

    el = driver.find_element(By.XPATH, "//span[text()='Select business entity']")
    driver.execute_script("arguments[0].scrollIntoView(true);", el)
    driver.find_element(By.XPATH, "//span[text()='Select business entity']").click()
    time.sleep(2)
    driver.find_element(By.XPATH,f"//span[text()='{bus_entity}']").click()
    time.sleep(1)
    
    element = driver.find_element(By.XPATH, "//span[text()='Identifier type']")
    driver.execute_script("arguments[0].scrollIntoView(true);", element)
    time.sleep(1)
    
    driver.find_element(By.XPATH, "//span[text()='Identifier type']").click()
    time.sleep(1)
    driver.find_element(By.XPATH, f"//span[text()='{loc_identifier_ty}']").click()
    time.sleep(1)

    identifier = driver.find_element(By.NAME,"locationNumber")
    identifier.clear()
    identifier.send_keys(loc_number)
    time.sleep(1)
    
    driver.find_element(By.XPATH,"//input[@id='extensionDigit']").click()
    time.sleep(1)

    driver.find_element(By.XPATH, "//button[text()='Next']").click()
    print("clicked on next button")
    time.sleep(3)

    try:
        next_page = EC.element_to_be_clickable((By.XPATH,"//span[text()='Select country']"))
        WebDriverWait(driver,10).until(next_page)
    except TimeoutException:
        print("Create location page: Timed out waiting for page to load")
    
    # Fill in the address details
    driver.find_element(By.XPATH, "//span[text()='Select country']").click()
    time.sleep(2)
    driver.find_element(By.XPATH,f"(//span[text()='{str(countryyy)}'])[2]").click()
    time.sleep(1)

    driver.find_element(By.XPATH, "//input[@name='state']").send_keys(state)
    driver.find_element(By.XPATH, "//input[@name='city']").send_keys(city)
    driver.find_element(By.XPATH, "//input[@name='address']").send_keys(address)
    driver.find_element(By.XPATH, "//input[@name='postalCode']").send_keys(postal_code)
    time.sleep(1)

    element = driver.find_element(By.XPATH, "//input[@name='contactPersonName']")
    driver.execute_script("arguments[0].scrollIntoView(true);", element)
    time.sleep(1)

    driver.find_element(By.XPATH, "//input[@placeholder='Enter name']").send_keys(contact_person)
    driver.find_element(By.XPATH, "//input[@placeholder='Enter email']").send_keys(email_id)
    time.sleep(1)

    driver.find_element(By.XPATH,"//button[@title='India']").click()
    time.sleep(1)
    element = driver.find_element(By.XPATH, "//span[text()='"+str(countryyy)+"']")
    driver.execute_script("arguments[0].scrollIntoView(true);", element)
    driver.find_element(By.XPATH,f"//span[text()='{countryyy}']").click()
    time.sleep(1)

    driver.find_element(By.ID,"custom-phone-input").send_keys(str(phone_number))
    driver.find_element(By.XPATH, "//input[@placeholder='Enter website']").send_keys(website)
    time.sleep(1)

    # Submit the form
    submit = driver.find_element(By.XPATH, "//button[text()='Submit']")
    submit.click()
    time.sleep(3)

    message = driver.find_element(By.XPATH,"//div[@class='p-toast-detail']").get_attribute("innerText")
    print(message)
    
    
    if "Successfully" in message:
        elm = EC.element_to_be_clickable((By.XPATH,"//span[text()= '+ New Location']"))
        WebDriverWait(driver,5).until(elm)
    time.sleep(2)
    
    sheet.cell(row=i, column=21).value=message
    workbook.save("ACG_Common_Workbook.xlsx")
