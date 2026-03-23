from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager
import tempfile

wb = load_workbook("ACG_Common_Workbook.xlsx",data_only=True)
typeev = wb["URL_Login_cred_Tenant"]
type_env = typeev.cell(row = 2, column = 4).value

options = Options()
options.add_argument(f'--user-data-dir={tempfile.mkdtemp()}')
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.implicitly_wait(60)


if type_env == "ACG":
    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_ACG"]

if type_env == "Tenant":
    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_Tenant"]

url = url_login_cred.cell(row = 2, column = 1).value
driver.get(str(url))

#driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[3]/i").click()

usern = url_login_cred.cell(row = 2, column = 2).value
driver.find_element(By.NAME,'identifier').send_keys(str(usern))

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))

driver.find_element(By.XPATH, "//button[text()='Login']").click()


roles_sheet = wb['Rights_test_case']
start_role = 8
last_role = 10
#(role.max_row) +1

#if type_env == 'ACG':

driver.find_element(By.XPATH,"//button[@aria-label='Menu']").click()
    
driver.find_element(By.XPATH,"//span[text()='Role Management']").click()
driver.find_element(By.XPATH,"//li[text()='View Roles']").click()



for i in range(start_role, last_role):

    timeout = 10
    try:
        new_role = EC.element_to_be_clickable((By.XPATH,"//span[text()='+ New Role']"))
        WebDriverWait(driver,timeout).until(new_role)
    except TimeoutException:
        print("Create role page: Timed out waiting for page to load")
    
    time.sleep(5)
    
    driver.find_element(By.XPATH,"//span[text()='+ New Role']").click()
    
    
    timeout = 10
    try:
        role_sub = EC.element_to_be_clickable((By.XPATH,"//span[text()='Submit']"))
        WebDriverWait(driver,timeout).until(role_sub)
    except TimeoutException:
        print("Create role page_submit: Timed out waiting for page to load")
    
    
    r_name = roles_sheet.cell(row = i, column = 3).value
    val = str(datetime.now()).split(".")
    ext=val[0].replace("-","").replace(":","").replace(" ","")
    driver.find_element(By.NAME,"roleName").send_keys(str(r_name)+ext)
    
    role_desp = roles_sheet.cell(row = i, column = 4).value
    driver.find_element(By.NAME,"roleDescription").send_keys(str(role_desp)+ext)
    
    time.sleep(2)
    
    el = driver.find_element(By.XPATH,"//label[text()='Select All']")
    driver.execute_script("arguments[0].scrollIntoView(true)",el)
    
    driver.find_element(By.XPATH,"//label[text()='Select All']").click()

    '''
    time.sleep(5)

    # = driver.find_elements(By.XPATH,"//div[@class='w-full']")
    roles = driver.find_elements(By.XPATH,"//span[@class='ml-2 text-base font-medium']")
    #print(roles)
    #for checkbox in roles:
            #print(checkbox)
    #        role = checkbox.text
    #        print("'"+role+"',")
            

    #//input[contains(@class,"chakra-checkbox__input")]
    if i == 8:
        
        for checkbox in roles:
            role = checkbox.text
            print(role)
            
            if role in ('Get All Events Based on EPC','Get All Events Based on Location Identifier',
                        'Get All Events Tenant','Get EPICS File Tenant','Get Single Event Epcis2',
                        'Get Events by ID','Upload Files to Data Lake','ViewSingleEPCISFile',
                        'Configure SFTP User Tenant','Get All Integrations','View Single Integration',
                        'Create Data Exchange Integrations','View All Data Exchange Integrations',
                        'Get file templates','View Templates','Create SNG Template',
                        'View Single Template','Update SNG Template','Update SSCC Template',
                        'Get SSCC Template Details By Id','Create SSCC Template',
                        'View All SSCC Templates','Get SNG History','Get SSCC Gen History',
                        'SSCC Generation','View Single SNG History Details',
                        'SSCC Generation Report','Execute L3 Serial Number Request',
                        'Execute L3 SSCC Request','Activate/Deactivate SNG Template',
                        'Activate/Deactivate SSCC Template'):
                #checkbox1 = driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div/div["+str(labell)+"]/div/label["+str(ide)+"]/span[2]")
                checkbox1 = driver.find_element(By.XPATH,f"//span[text()='{role}']")
                driver.execute_script("arguments[0].scrollIntoView()", checkbox1)
                driver.find_element(By.XPATH,f"//span[text()='{role}']").click()

    if i == 9:
        
        for checkbox in roles:
            role = checkbox.text
            print(role)
            
            if role in ('View Templates','Create SNG Template',
                        'View Single Template','Update SNG Template','Update SSCC Template',
                        'Get SSCC Template Details By Id','Create SSCC Template',
                        'View All SSCC Templates','Get SNG History','Get SSCC Gen History',
                        'SSCC Generation','View Single SNG History Details',
                        'SSCC Generation Report','Execute L3 Serial Number Request',
                        'Execute L3 SSCC Request','Activate/Deactivate SNG Template',
                        'Activate/Deactivate SSCC Template'):
            
                checkbox1 = driver.find_element(By.XPATH,f"//span[text()='{role}']")
                driver.execute_script("arguments[0].scrollIntoView()", checkbox1)
                driver.find_element(By.XPATH,f"//span[text()='{role}']").click()
    
    '''
    el2 = driver.find_element(By.XPATH,"//span[text()='Submit']")
    driver.execute_script("arguments[0].scrollIntoView(true)",el2)
    driver.find_element(By.XPATH,"//span[text()='Submit']").click()
    message = driver.find_element(By.XPATH, "//div[@class='p-toast-detail']").get_attribute("innerText")
    print(message)
    #waiting for new role button
    if "created" in message:
        print(f"Role {r_name+ext} Created successfully")
    else:
        #click on cancel button
        print(f"Role {r_name+ext} already exists")


    roles_sheet.cell(row = i, column = 5).value = message
    wb.save("ACG_Common_Workbook.xlsx")